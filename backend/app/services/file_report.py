from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path

import duckdb
from openpyxl import Workbook

from app.parsers.csv_writer import CSV_DELIMITER, duckdb_read_csv_delim
from app.paths import get_report_output_dir, load_report_query_sql
from app.services.csv_storage import CSV_TO_DB, list_all_csv_paths, resolve_csv_paths_for_dates
from app.services.device_detection import parse_browser_device
from app.services.report import (
    DEFAULT_LIMIT,
    MAX_LIMIT,
    ReportResult,
    format_report_cell_value,
    normalize_date_from,
    normalize_date_to,
    validate_custom_sql,
)

DATE_ONLY_PATTERN = re.compile(r"^\d{4}-\d{2}-\d{2}$")
SQL_FILTER_DATE_PATTERN = re.compile(
    r"areq\.messagedatetime\s*>=\s*'(\d{4}-\d{2}-\d{2})",
    re.IGNORECASE,
)


def _iter_dates(date_from: str, date_to: str | None) -> list[str]:
    from_day = normalize_date_from(date_from)[:10]
    normalized_to = normalize_date_to(date_to or date_from) or normalize_date_from(date_from)
    to_day = normalized_to[:10]
    start = date.fromisoformat(from_day)
    end = date.fromisoformat(to_day)
    if end < start:
        raise ValueError("dateTo must be on or after dateFrom.")

    dates: list[str] = []
    current = start
    while current <= end:
        dates.append(current.isoformat())
        current += timedelta(days=1)
    return dates


def _extract_sql_filter_dates(sql: str) -> list[str]:
    return sorted({match.group(1) for match in SQL_FILTER_DATE_PATTERN.finditer(sql)})


def _resolve_csv_paths_for_report(
    *,
    mode: str,
    date_from: str | None,
    date_to: str | None,
    sql: str | None,
) -> list[Path]:
    dates: list[str] = []
    if date_from and date_from.strip():
        dates = _iter_dates(date_from, date_to)

    if mode == "custom" and sql:
        sql_dates = _extract_sql_filter_dates(sql)
        if sql_dates:
            dates = sorted(set(dates) | set(sql_dates))

    if dates:
        return resolve_csv_paths_for_dates(dates)
    return list_all_csv_paths()


def _csv_header_columns(csv_paths: list[Path]) -> set[str]:
    if not csv_paths:
        return set()
    columns: set[str] = set()
    for path in csv_paths:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle, delimiter=CSV_DELIMITER)
            columns.update(next(reader, []))
    return columns


def _register_browser_device_functions(connection: duckdb.DuckDBPyConnection) -> None:
    def browser_os_from_ua(user_agent: str | None) -> str:
        return parse_browser_device(user_agent).os

    def browser_model_from_ua(user_agent: str | None) -> str:
        return parse_browser_device(user_agent).model

    connection.create_function("vst_browser_os", browser_os_from_ua)
    connection.create_function("vst_browser_model", browser_model_from_ua)


def _duckdb_column_expr(csv_col: str, db_col: str, available: set[str]) -> str:
    if csv_col in available:
        return f'"{csv_col}" AS {db_col}'
    return f"NULL::VARCHAR AS {db_col}"


def _duckdb_messages_table_sql(csv_paths: list[Path]) -> str:
    paths_literal = ", ".join(f"'{path.as_posix()}'" for path in csv_paths)
    available = _csv_header_columns(csv_paths)
    select_columns = ",\n        ".join(
        _duckdb_column_expr(csv_col, db_col, available)
        for csv_col, db_col in CSV_TO_DB.items()
    )
    return f"""
        CREATE OR REPLACE TABLE cust_acs_3dsmess AS
        SELECT
            {select_columns}
        FROM read_csv([{paths_literal}], header=true, delim='{duckdb_read_csv_delim()}', union_by_name=true, all_varchar=true)
    """


def _enrich_browser_device_columns(connection: duckdb.DuckDBPyConnection) -> None:
    columns = {
        row[0]
        for row in connection.execute(
            "SELECT column_name FROM information_schema.columns "
            "WHERE table_name = 'cust_acs_3dsmess'"
        ).fetchall()
    }
    if "browseruseragent" not in columns:
        return
    _register_browser_device_functions(connection)
    connection.execute(
        """
        CREATE OR REPLACE TEMP TABLE ua_device_lookup AS
        SELECT DISTINCT
            browseruseragent,
            vst_browser_os(browseruseragent) AS browseros,
            vst_browser_model(browseruseragent) AS browsermodel
        FROM cust_acs_3dsmess
        WHERE browseruseragent IS NOT NULL AND browseruseragent != ''
        """
    )
    if "browseros" not in columns:
        connection.execute(
            "ALTER TABLE cust_acs_3dsmess ADD COLUMN browseros VARCHAR"
        )
    if "browsermodel" not in columns:
        connection.execute(
            "ALTER TABLE cust_acs_3dsmess ADD COLUMN browsermodel VARCHAR"
        )
    connection.execute(
        """
        UPDATE cust_acs_3dsmess
        SET
            browseros = COALESCE(NULLIF(l.browseros, ''), NULLIF(cust_acs_3dsmess.browseros, '')),
            browsermodel = COALESCE(
                NULLIF(l.browsermodel, ''), NULLIF(cust_acs_3dsmess.browsermodel, '')
            )
        FROM ua_device_lookup l
        WHERE cust_acs_3dsmess.browseruseragent = l.browseruseragent
        """
    )


def _csv_file_signature(csv_paths: list[Path]) -> tuple:
    return tuple(
        (path.as_posix(), path.stat().st_mtime_ns, path.stat().st_size)
        for path in csv_paths
    )


_excel_ready_connections: set[int] = set()


def _ensure_duckdb_excel(connection: duckdb.DuckDBPyConnection) -> bool:
    connection_id = id(connection)
    if connection_id in _excel_ready_connections:
        return True
    try:
        connection.execute("INSTALL excel")
        connection.execute("LOAD excel")
        _excel_ready_connections.add(connection_id)
        return True
    except Exception:
        return False


def _adapt_report_sql_for_duckdb(sql: str) -> str:
    adapted = sql.replace("%%", "%")
    adapted = adapted.replace("%(txn_id)s::text IS NULL", "(? IS NULL)")
    for table in ("areq", "ds"):
        adapted = adapted.replace(
            f"{table}.threedsservertransid = %(txn_id)s::text",
            f"{table}.threedsservertransid = ?",
        )
        adapted = adapted.replace(
            f"{table}.messagedatetime >= %(date_from)s::text",
            f"{table}.messagedatetime >= ?",
        )
        adapted = adapted.replace(
            f"(%(date_to)s::text IS NULL OR {table}.messagedatetime <= %(date_to)s::text)",
            f"(? IS NULL OR {table}.messagedatetime <= ?)",
        )
    return adapted


def _report_params(date_from: str, date_to: str | None, txn_id: str | None) -> tuple:
    return (
        txn_id,
        txn_id,
        date_from,
        date_to,
        date_to,
    )


def _build_report_output_name(
    *,
    mode: str,
    date_from: str | None,
    date_to: str | None,
    txn_id: str | None,
) -> str:
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S-%f")
    if mode == "txnId" and txn_id:
        safe_txn = re.sub(r"[^0-9a-fA-F-]", "", txn_id.strip())
        return f"report-txn-{safe_txn}-{stamp}.xlsx"
    from_day = (date_from or "unknown")[:10]
    to_day = (date_to or date_from or from_day)[:10]
    if from_day == to_day:
        return f"report-{from_day}-{stamp}.xlsx"
    return f"report-{from_day}-to-{to_day}-{stamp}.xlsx"


PIVOT_ROW_FIELD = "txn_result"
PIVOT_VALUE_FIELD = "threedsservertransid"
NATIVE_PIVOT_MAX_ROWS = 5_000
SUMMARY_APPEND_MAX_ROWS = 25_000
PIVOT_ROW_FIELDS = [
    "r02",
    "areq_messagedate",
    "oob_missing_day",
    "final_cres_status",
    "txn_timeline",
    "browser_user_agent",
    "merchant_name",
    "browser_os",
    "browser_model",
    "threedsservertransid",
    "oob_missing",
]


def _compute_summary(
    connection: duckdb.DuckDBPyConnection,
    columns: list[str],
    total: int,
) -> list[tuple[str, int, float]]:
    if PIVOT_ROW_FIELD not in columns or total <= 0:
        return []
    rows = connection.execute(
        f'SELECT COALESCE("{PIVOT_ROW_FIELD}", \'(none)\') AS k, COUNT(*) AS c '
        "FROM report_result GROUP BY 1 ORDER BY c DESC"
    ).fetchall()
    return [(str(key), int(count), count / total) for key, count in rows]


def _append_summary_sheet(
    workbook,
    summary: list[tuple[str, int, float]],
) -> None:
    from openpyxl.cell import WriteOnlyCell

    sheet = workbook.create_sheet("Summary")
    sheet.append([PIVOT_ROW_FIELD, "count", "percent"])
    for key, count, percent in summary:
        percent_cell = WriteOnlyCell(sheet, value=percent)
        percent_cell.number_format = "0.0%"
        sheet.append([key, count, percent_cell])


def _append_summary_to_existing_xlsx(
    output_path: Path,
    summary: list[tuple[str, int, float]],
) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(output_path)
    if "Summary" in workbook.sheetnames:
        del workbook["Summary"]
    _append_summary_sheet(workbook, summary)
    workbook.save(output_path)


def _save_report_xlsx_openpyxl(
    connection: duckdb.DuckDBPyConnection,
    columns: list[str],
    output_path: Path,
    summary: list[tuple[str, int, float]],
) -> None:
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Data")
    sheet.append(columns)
    cursor = connection.execute("SELECT * FROM report_result ORDER BY 1")
    while True:
        batch = cursor.fetchmany(5000)
        if not batch:
            break
        for row in batch:
            sheet.append(
                [
                    ""
                    if value is None
                    else format_report_cell_value(columns[index], str(value))
                    for index, value in enumerate(row)
                ]
            )
    if summary:
        _append_summary_sheet(workbook, summary)
    workbook.save(output_path)


def _save_report_xlsx_duckdb(
    connection: duckdb.DuckDBPyConnection,
    output_path: Path,
    summary: list[tuple[str, int, float]],
    total: int,
) -> bool:
    if not _ensure_duckdb_excel(connection):
        return False
    path = str(output_path.resolve()).replace("'", "''")
    copy_sql = (
        f"COPY (SELECT * FROM report_result ORDER BY 1) TO '{path}' "
        "WITH (FORMAT xlsx, HEADER true"
    )
    try:
        connection.execute(f"{copy_sql}, SHEET 'Data')")
    except Exception:
        try:
            connection.execute(f"{copy_sql})")
            from openpyxl import load_workbook

            workbook = load_workbook(output_path)
            workbook.active.title = "Data"
            workbook.save(output_path)
        except Exception:
            return False
    if summary and total <= SUMMARY_APPEND_MAX_ROWS:
        _append_summary_to_existing_xlsx(output_path, summary)
    return True


def _save_report_xlsx(
    connection: duckdb.DuckDBPyConnection,
    columns: list[str],
    output_path: Path,
    summary: list[tuple[str, int, float]],
    total: int,
) -> None:
    if _save_report_xlsx_duckdb(connection, output_path, summary, total):
        return
    _save_report_xlsx_openpyxl(connection, columns, output_path, summary)


def _can_build_pivot(columns: list[str], total: int) -> bool:
    if total <= 0 or PIVOT_VALUE_FIELD not in columns:
        return False
    return all(field in columns for field in PIVOT_ROW_FIELDS)


def _add_refresh_on_open_pivot(
    output_path: Path,
    columns: list[str],
    total: int,
) -> bool:
    from openpyxl import load_workbook

    from app.services.excel_pivot import add_pivot_sheet

    workbook = load_workbook(output_path)
    added = add_pivot_sheet(
        workbook,
        columns=columns,
        total=total,
        row_fields=PIVOT_ROW_FIELDS,
        value_field=PIVOT_VALUE_FIELD,
    )
    if added:
        workbook.save(output_path)
    return added


def _add_report_pivot(
    output_path: Path,
    columns: list[str],
    total: int,
    *,
    native_pivot: bool,
) -> tuple[bool, str]:
    if not native_pivot:
        return False, ""

    from app.services.excel_pivot import add_native_pivot, native_pivot_available

    if total <= NATIVE_PIVOT_MAX_ROWS and native_pivot_available():
        try:
            add_native_pivot(
                output_path,
                data_sheet="Data",
                data_rows=total,
                data_cols=len(columns),
                row_fields=PIVOT_ROW_FIELDS,
                value_field=PIVOT_VALUE_FIELD,
            )
            return True, ""
        except Exception as error:
            if _add_refresh_on_open_pivot(output_path, columns, total):
                return True, (
                    "Excel automation failed, added a refresh-on-open pivot instead "
                    f"(open in Excel to populate it): {error}"
                )
            return False, str(error)

    if _add_refresh_on_open_pivot(output_path, columns, total):
        return True, (
            "Added a refresh-on-open pivot. Open the file in Excel to populate it."
        )
    return False, "Could not build the pivot sheet."


@dataclass
class ReportExportResult:
    columns: list[str]
    row_count: int
    file_name: str
    output_path: Path
    pivot_added: bool = False
    pivot_error: str = ""


def _sql_has_report_placeholders(sql: str) -> bool:
    return (
        "%(date_from)s" in sql
        or "%(date_to)s" in sql
        or "%(txn_id)s" in sql
    )


def _resolve_report_query(
    *,
    mode: str,
    date_from: str | None,
    date_to: str | None,
    txn_id: str | None,
    sql: str | None,
) -> tuple[list[Path], str, tuple]:
    if mode == "custom":
        if not sql:
            raise ValueError("Custom mode requires sql.")
        validate_custom_sql(sql)
        csv_paths = _resolve_csv_paths_for_report(
            mode=mode,
            date_from=date_from,
            date_to=date_to,
            sql=sql,
        )
        report_sql = sql.strip().rstrip(";")
        if _sql_has_report_placeholders(report_sql):
            normalized_from = normalize_date_from(date_from or "1970-01-01")
            normalized_to = normalize_date_to(date_to)
            report_sql = _adapt_report_sql_for_duckdb(report_sql)
            bound_txn = txn_id.strip() if txn_id and txn_id.strip() else None
            params = _report_params(normalized_from, normalized_to, bound_txn)
        else:
            params = ()
    elif mode == "txnId":
        if not txn_id or not txn_id.strip():
            raise ValueError("Transaction ID is required.")
        normalized_from = normalize_date_from(date_from or "1970-01-01")
        normalized_to = normalize_date_to(date_to)
        csv_paths = _resolve_csv_paths_for_report(
            mode=mode,
            date_from=normalized_from if date_from and date_from.strip() else None,
            date_to=normalized_to,
            sql=None,
        )
        report_sql = _adapt_report_sql_for_duckdb(load_report_query_sql())
        params = _report_params(normalized_from, normalized_to, txn_id.strip())
    elif mode == "date":
        if not date_from or not date_from.strip():
            raise ValueError("dateFrom is required.")
        normalized_from = normalize_date_from(date_from)
        normalized_to = normalize_date_to(date_to)
        csv_paths = _resolve_csv_paths_for_report(
            mode=mode,
            date_from=normalized_from,
            date_to=normalized_to,
            sql=None,
        )
        report_sql = _adapt_report_sql_for_duckdb(load_report_query_sql())
        params = _report_params(normalized_from, normalized_to, None)
    else:
        raise ValueError(f"Unsupported mode: {mode}")

    return csv_paths, report_sql, params


def _load_report_result(
    connection: duckdb.DuckDBPyConnection,
    csv_paths: list[Path],
    report_sql: str,
    params: tuple,
    *,
    reload_messages: bool,
) -> list[str]:
    if reload_messages:
        connection.execute(_duckdb_messages_table_sql(csv_paths))
        _enrich_browser_device_columns(connection)
    connection.execute("DROP TABLE IF EXISTS report_result")
    connection.execute(f"CREATE TEMP TABLE report_result AS {report_sql}", params)
    return [
        row[0]
        for row in connection.execute(
            "SELECT column_name FROM information_schema.columns "
            "WHERE table_name = 'report_result' ORDER BY ordinal_position"
        ).fetchall()
    ]


@dataclass
class _CachedReport:
    csv_signature: tuple
    report_signature: tuple
    connection: duckdb.DuckDBPyConnection
    columns: list[str]
    total: int


_cache: _CachedReport | None = None


def clear_report_cache() -> None:
    global _cache
    if _cache is not None:
        try:
            _excel_ready_connections.discard(id(_cache.connection))
            _cache.connection.close()
        except Exception:
            pass
        _cache = None


def _report_signature(
    *,
    mode: str,
    report_sql: str,
    params: tuple,
) -> tuple:
    return (mode, report_sql, params)


def _materialized_report(
    *,
    mode: str,
    csv_paths: list[Path],
    report_sql: str,
    params: tuple,
) -> _CachedReport:
    global _cache
    csv_signature = _csv_file_signature(csv_paths)
    report_signature = _report_signature(mode=mode, report_sql=report_sql, params=params)
    if _cache is not None:
        if (
            _cache.csv_signature == csv_signature
            and _cache.report_signature == report_signature
        ):
            return _cache
        if _cache.csv_signature == csv_signature:
            connection = _cache.connection
            try:
                columns = _load_report_result(
                    connection,
                    csv_paths,
                    report_sql,
                    params,
                    reload_messages=False,
                )
                count_row = connection.execute("SELECT COUNT(*) FROM report_result").fetchone()
                total = int(count_row[0]) if count_row else 0
            except Exception:
                clear_report_cache()
            else:
                _cache = _CachedReport(
                    csv_signature=csv_signature,
                    report_signature=report_signature,
                    connection=connection,
                    columns=columns,
                    total=total,
                )
                return _cache

    clear_report_cache()
    connection = duckdb.connect()
    try:
        columns = _load_report_result(
            connection,
            csv_paths,
            report_sql,
            params,
            reload_messages=True,
        )
        count_row = connection.execute("SELECT COUNT(*) FROM report_result").fetchone()
        total = int(count_row[0]) if count_row else 0
    except Exception:
        connection.close()
        raise
    _cache = _CachedReport(
        csv_signature=csv_signature,
        report_signature=report_signature,
        connection=connection,
        columns=columns,
        total=total,
    )
    return _cache


def run_report_query(
    *,
    mode: str,
    date_from: str | None = None,
    date_to: str | None = None,
    txn_id: str | None = None,
    sql: str | None = None,
    limit: int = DEFAULT_LIMIT,
    offset: int = 0,
) -> ReportResult:
    limit = min(max(limit, 1), MAX_LIMIT)
    offset = max(offset, 0)

    csv_paths, report_sql, params = _resolve_report_query(
        mode=mode,
        date_from=date_from,
        date_to=date_to,
        txn_id=txn_id,
        sql=sql,
    )

    cached = _materialized_report(
        mode=mode,
        csv_paths=csv_paths,
        report_sql=report_sql,
        params=params,
    )
    page_rows_raw = cached.connection.execute(
        "SELECT * FROM report_result ORDER BY 1 LIMIT ? OFFSET ?",
        [limit, offset],
    ).fetchall()
    page_rows = [
        {col: ("" if value is None else str(value)) for col, value in zip(cached.columns, row)}
        for row in page_rows_raw
    ]
    return ReportResult(
        columns=cached.columns,
        rows=page_rows,
        row_count=cached.total,
        limit=limit,
        offset=offset,
    )


def export_report_xlsx(
    *,
    mode: str,
    date_from: str | None = None,
    date_to: str | None = None,
    txn_id: str | None = None,
    sql: str | None = None,
    native_pivot: bool = False,
) -> ReportExportResult:
    csv_paths, report_sql, params = _resolve_report_query(
        mode=mode,
        date_from=date_from,
        date_to=date_to,
        txn_id=txn_id,
        sql=sql,
    )
    output_name = _build_report_output_name(
        mode=mode,
        date_from=date_from,
        date_to=date_to,
        txn_id=txn_id,
    )
    output_path = get_report_output_dir() / output_name
    output_path.parent.mkdir(parents=True, exist_ok=True)

    cached = _materialized_report(
        mode=mode,
        csv_paths=csv_paths,
        report_sql=report_sql,
        params=params,
    )
    summary = _compute_summary(cached.connection, cached.columns, cached.total)
    _save_report_xlsx(cached.connection, cached.columns, output_path, summary, cached.total)

    pivot_added = False
    pivot_error = ""
    if native_pivot and _can_build_pivot(cached.columns, cached.total):
        pivot_added, pivot_error = _add_report_pivot(
            output_path,
            cached.columns,
            cached.total,
            native_pivot=native_pivot,
        )

    return ReportExportResult(
        columns=cached.columns,
        row_count=cached.total,
        file_name=output_name,
        output_path=output_path,
        pivot_added=pivot_added,
        pivot_error=pivot_error,
    )
