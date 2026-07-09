from __future__ import annotations

from pathlib import Path

import duckdb

from app.services.report import normalize_date_from, normalize_date_to

CHAMPIONS_SHEET = "Champions"
CHAMPION_ATTEMPTS_SHEET = "Champion Attempts"

CHAMPIONS_COLUMNS = [
    "champion_rank",
    "day",
    "card_masked",
    "merchant_name",
    "merchant_key",
    "attempts",
    "first_attempt",
    "last_attempt",
    "results_mix",
]

CHAMPION_ATTEMPTS_COLUMNS = [
    "champion_rank",
    "card_masked",
    "merchant_name",
    "merchant_key",
    "attempt_no",
    "attempts_total",
    "areq_messagedatetime",
    "threedsservertransid",
    "txn_result",
    "ares_status",
]

_TXN_BASE_SQL = """
WITH merchant_lookup AS (
    SELECT
        threedsservertransid,
        max(NULLIF(trim(acquirermerchantid), '')) AS acquirer_merchant_id
    FROM cust_acs_3dsmess
    WHERE messagetype = 'AReq'
    GROUP BY threedsservertransid
),
txn AS (
    SELECT
        substr(r.areq_messagedatetime, 1, 10) AS day,
        CASE
            WHEN r.acct_number IS NULL OR trim(r.acct_number) = '' THEN '(no card)'
            WHEN length(trim(r.acct_number)) >= 10
            THEN left(trim(r.acct_number), 6) || '******' || right(trim(r.acct_number), 4)
            ELSE trim(r.acct_number)
        END AS card_masked,
        COALESCE(
            NULLIF(trim(m.acquirer_merchant_id), ''),
            NULLIF(trim(r.merchant_name), ''),
            '(unknown)'
        ) AS merchant_key,
        COALESCE(NULLIF(trim(r.merchant_name), ''), '(unknown)') AS merchant_name,
        r.areq_messagedatetime,
        r.threedsservertransid,
        COALESCE(r.txn_result, '(none)') AS txn_result,
        r.ares_status
    FROM report_result r
    LEFT JOIN merchant_lookup m ON r.threedsservertransid = m.threedsservertransid
),
grouped AS (
    SELECT
        day,
        card_masked,
        merchant_key,
        max(merchant_name) AS merchant_name,
        count(*) AS attempts,
        min(areq_messagedatetime) AS first_attempt,
        max(areq_messagedatetime) AS last_attempt,
        string_agg(DISTINCT txn_result, ', ') AS results_mix
    FROM txn
    GROUP BY day, card_masked, merchant_key
),
ranked AS (
    SELECT
        row_number() OVER (
            ORDER BY attempts DESC, last_attempt DESC, card_masked, merchant_key
        ) AS champion_rank,
        day,
        card_masked,
        merchant_key,
        merchant_name,
        attempts,
        first_attempt,
        last_attempt,
        results_mix
    FROM grouped
)
"""

_CHAMPIONS_SQL = f"""
{_TXN_BASE_SQL}
SELECT
    champion_rank,
    day,
    card_masked,
    merchant_name,
    merchant_key,
    attempts,
    first_attempt,
    last_attempt,
    results_mix
FROM ranked
ORDER BY champion_rank
"""

_CHAMPION_ATTEMPTS_SQL = f"""
{_TXN_BASE_SQL}
SELECT
    rk.champion_rank,
    t.card_masked,
    t.merchant_name,
    t.merchant_key,
    row_number() OVER (
        PARTITION BY t.day, t.card_masked, t.merchant_key
        ORDER BY t.areq_messagedatetime, t.threedsservertransid
    ) AS attempt_no,
    rk.attempts AS attempts_total,
    t.areq_messagedatetime,
    t.threedsservertransid,
    t.txn_result,
    t.ares_status
FROM txn t
INNER JOIN ranked rk
    ON t.day = rk.day
   AND t.card_masked = rk.card_masked
   AND t.merchant_key = rk.merchant_key
ORDER BY rk.champion_rank, attempt_no
"""


def mask_card_number(acct_number: str | None) -> str:
    value = (acct_number or "").strip()
    if not value:
        return "(no card)"
    if len(value) >= 10:
        return f"{value[:6]}******{value[-4:]}"
    return value


def is_single_day_date_report(
    *,
    mode: str,
    date_from: str | None,
    date_to: str | None,
) -> bool:
    if mode != "date" or not date_from or not date_from.strip():
        return False
    from_day = normalize_date_from(date_from)[:10]
    if not date_to or not date_to.strip():
        return True
    to_day = normalize_date_to(date_to)[:10]
    return from_day == to_day


def _required_report_columns(connection: duckdb.DuckDBPyConnection) -> set[str]:
    rows = connection.execute(
        "SELECT column_name FROM information_schema.columns "
        "WHERE table_name = 'report_result'"
    ).fetchall()
    return {row[0] for row in rows}


def can_build_champion_sheets(connection: duckdb.DuckDBPyConnection) -> bool:
    columns = _required_report_columns(connection)
    required = {
        "acct_number",
        "merchant_name",
        "areq_messagedatetime",
        "threedsservertransid",
        "txn_result",
        "ares_status",
    }
    if not required.issubset(columns):
        return False
    try:
        connection.execute("SELECT 1 FROM cust_acs_3dsmess LIMIT 1")
    except Exception:
        return False
    return True


def fetch_champions_rows(
    connection: duckdb.DuckDBPyConnection,
) -> list[tuple]:
    return connection.execute(_CHAMPIONS_SQL).fetchall()


def fetch_champion_attempt_rows(
    connection: duckdb.DuckDBPyConnection,
) -> list[tuple]:
    return connection.execute(_CHAMPION_ATTEMPTS_SQL).fetchall()


def _write_sheet(workbook, sheet_name: str, columns: list[str], rows: list[tuple]) -> None:
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    sheet = workbook.create_sheet(sheet_name)
    sheet.append(columns)
    for row in rows:
        sheet.append(["" if value is None else value for value in row])


def append_champion_sheets(
    connection: duckdb.DuckDBPyConnection,
    output_path: Path,
) -> bool:
    if not can_build_champion_sheets(connection):
        return False
    champions = fetch_champions_rows(connection)
    attempts = fetch_champion_attempt_rows(connection)
    from openpyxl import load_workbook

    workbook = load_workbook(output_path)
    _write_sheet(workbook, CHAMPIONS_SHEET, CHAMPIONS_COLUMNS, champions)
    _write_sheet(workbook, CHAMPION_ATTEMPTS_SHEET, CHAMPION_ATTEMPTS_COLUMNS, attempts)
    workbook.save(output_path)
    return True
