import os
import tempfile
import unittest
from pathlib import Path

from app.parsers.models import MessageRow
from app.services import file_report
from app.services.csv_storage import resolve_csv_paths_for_dates, save_daily_csvs, CSV_TO_DB


def _areq(txn: str, when: str, *, acct: str = "4111111111111111") -> MessageRow:
    return MessageRow(
        log_file="elastic.log",
        message_datetime=when,
        message_type="AReq",
        three_ds_server_trans_id=txn,
        acct_number=acct,
        merchant_name="Test Merchant",
        browser_user_agent="Mozilla/5.0 (Linux; Android 14; SM-S921B) AppleWebKit/537.36",
    )


def _ares(txn: str, when: str, *, status: str = "C", reason: str = "") -> MessageRow:
    return MessageRow(
        log_file="elastic.log",
        message_datetime=when,
        message_type="ARes",
        three_ds_server_trans_id=txn,
        trans_status=status,
        trans_status_reason=reason,
    )


def _cres(txn: str, when: str, *, status: str = "Y") -> MessageRow:
    return MessageRow(
        log_file="elastic.log",
        message_datetime=when,
        message_type="CRes",
        three_ds_server_trans_id=txn,
        trans_status=status,
    )


class FileReportTest(unittest.TestCase):
    def test_pivot_row_fields_order(self) -> None:
        self.assertEqual(
            file_report.PIVOT_ROW_FIELDS,
            [
                "r02",
                "areq_messagedate",
                "browser_os",
                "browser_model",
                "oob_missing_day",
                "final_cres_status",
                "txn_timeline",
                "browser_user_agent",
                "merchant_name",
                "three_ds_requestor_info",
                "threedsservertransid",
                "oob_missing",
            ],
        )

    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self._saved = {
            key: os.environ.get(key)
            for key in ("CSV_STORAGE_DIR", "REPORT_OUTPUT_DIR")
        }
        os.environ["CSV_STORAGE_DIR"] = str(Path(self._tmp.name) / "csv")
        os.environ["REPORT_OUTPUT_DIR"] = str(Path(self._tmp.name) / "reports")

    def tearDown(self) -> None:
        for key, value in self._saved.items():
            if value is None:
                os.environ.pop(key, None)
            else:
                os.environ[key] = value
        self._tmp.cleanup()

    def _seed_day(self, day: str, txn: str = "T1") -> None:
        save_daily_csvs(
            [
                _areq(txn, f"{day} 10:00:00.000"),
                _ares(txn, f"{day} 10:00:01.000"),
                _cres(txn, f"{day} 10:05:00.000"),
            ]
        )

    def test_default_report_single_day(self) -> None:
        self._seed_day("2026-06-20", "T1")
        result = file_report.run_report_query(
            mode="date",
            date_from="2026-06-20 00:00:00",
            date_to="2026-06-20 23:59:59",
        )
        self.assertEqual(result.row_count, 1)
        self.assertIn("threedsservertransid", result.columns)
        row = result.rows[0]
        self.assertEqual(row["threedsservertransid"], "T1")
        self.assertIn("AReq", row["txn_timeline"])
        self.assertIn("ARes(C+NULL)", row["txn_timeline"])
        self.assertIn("CRes(Y)", row["txn_timeline"])
        self.assertEqual(row["browser_os"], "Android")
        self.assertEqual(row["browser_model"], "Samsung SM-S921B")
        self.assertEqual(row["card_scheme"], "Visa")

    def test_report_fills_browser_fields_from_user_agent_without_stored_columns(self) -> None:
        from app.parsers.csv_writer import csv_dict_writer

        csv_dir = Path(os.environ["CSV_STORAGE_DIR"])
        csv_dir.mkdir(parents=True, exist_ok=True)
        legacy_columns = [
            column
            for column in CSV_TO_DB.keys()
            if column not in {"browserOS", "browserModel"}
        ]
        row = {
            "messageDateTime": "2026-06-20 10:00:00.000",
            "messageType": "AReq",
            "threeDSServerTransID": "T1",
            "acctNumber": "4111111111111111",
            "merchantName": "Test Merchant",
            "browserUserAgent": (
                "Mozilla/5.0 (Linux; Android 14; SM-S921B) AppleWebKit/537.36"
            ),
        }
        csv_path = csv_dir / "2026-06-20.csv"
        with csv_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv_dict_writer(handle, fieldnames=legacy_columns, extrasaction="ignore")
            writer.writeheader()
            writer.writerow(row)
        file_report.clear_report_cache()
        result = file_report.run_report_query(
            mode="date",
            date_from="2026-06-20 00:00:00",
            date_to="2026-06-20 23:59:59",
        )
        self.assertEqual(result.row_count, 1)
        self.assertEqual(result.rows[0]["browser_os"], "Android")
        self.assertEqual(result.rows[0]["browser_model"], "Samsung SM-S921B")

    def test_ares_timeline_includes_status_reason(self) -> None:
        save_daily_csvs(
            [
                _areq("T1", "2026-06-20 10:00:00.000"),
                _ares("T1", "2026-06-20 10:00:01.000", status="R", reason="02"),
            ]
        )
        result = file_report.run_report_query(
            mode="date",
            date_from="2026-06-20 00:00:00",
            date_to="2026-06-20 23:59:59",
        )
        self.assertIn("ARes(R+02)", result.rows[0]["txn_timeline"])
        self.assertEqual(result.rows[0]["r02"], "YES")

    def test_r02_is_no_without_ares_r02(self) -> None:
        self._seed_day("2026-06-20", "T1")
        result = file_report.run_report_query(
            mode="date",
            date_from="2026-06-20 00:00:00",
            date_to="2026-06-20 23:59:59",
        )
        self.assertEqual(result.rows[0]["r02"], "NO")

    def test_multi_day_report_counts_all(self) -> None:
        self._seed_day("2026-06-20", "T1")
        self._seed_day("2026-06-21", "T2")
        result = file_report.run_report_query(
            mode="date",
            date_from="2026-06-20 00:00:00",
            date_to="2026-06-21 23:59:59",
        )
        self.assertEqual(result.row_count, 2)
        txns = {row["threedsservertransid"] for row in result.rows}
        self.assertEqual(txns, {"T1", "T2"})

    def test_missing_csv_day_in_range_blocks(self) -> None:
        self._seed_day("2026-06-20", "T1")
        self._seed_day("2026-06-22", "T2")
        with self.assertRaises(ValueError) as ctx:
            file_report.run_report_query(
                mode="date",
                date_from="2026-06-20 00:00:00",
                date_to="2026-06-22 23:59:59",
            )
        self.assertIn("2026-06-21", str(ctx.exception))

    def test_partial_day_csv_is_accepted(self) -> None:
        # A partial day still produces a CSV file (only morning data).
        save_daily_csvs(
            [
                _areq("T1", "2026-06-20 07:00:00.000"),
                _ares("T1", "2026-06-20 07:00:01.000"),
            ]
        )
        result = file_report.run_report_query(
            mode="date",
            date_from="2026-06-20 00:00:00",
            date_to="2026-06-20 23:59:59",
        )
        self.assertEqual(result.row_count, 1)

    def test_txn_mode_returns_only_requested(self) -> None:
        save_daily_csvs(
            [
                _areq("T1", "2026-06-20 10:00:00.000"),
                _ares("T1", "2026-06-20 10:00:01.000"),
                _areq("T2", "2026-06-20 11:00:00.000"),
                _ares("T2", "2026-06-20 11:00:01.000"),
            ]
        )
        result = file_report.run_report_query(
            mode="txnId",
            date_from="2026-06-20 00:00:00",
            date_to="2026-06-20 23:59:59",
            txn_id="T2",
        )
        self.assertEqual(result.row_count, 1)
        self.assertEqual(result.rows[0]["threedsservertransid"], "T2")

    def test_export_writes_unique_filenames(self) -> None:
        self._seed_day("2026-06-20", "T1")
        first = file_report.export_report_xlsx(
            mode="date",
            date_from="2026-06-20 00:00:00",
            date_to="2026-06-20 23:59:59",
            native_pivot=False,
        )
        second = file_report.export_report_xlsx(
            mode="date",
            date_from="2026-06-20 00:00:00",
            date_to="2026-06-20 23:59:59",
            native_pivot=False,
        )
        self.assertTrue(Path(first.output_path).exists())
        self.assertTrue(Path(second.output_path).exists())
        self.assertNotEqual(first.file_name, second.file_name)
        self.assertEqual(first.row_count, 1)

    def test_export_has_data_and_summary_sheets(self) -> None:
        from openpyxl import load_workbook

        save_daily_csvs(
            [
                _areq("T1", "2026-06-20 10:00:00.000"),
                _ares("T1", "2026-06-20 10:00:01.000"),
                _cres("T1", "2026-06-20 10:05:00.000", status="Y"),
                _areq("T2", "2026-06-20 11:00:00.000"),
                _ares("T2", "2026-06-20 11:00:01.000"),
            ]
        )
        export = file_report.export_report_xlsx(
            mode="date",
            date_from="2026-06-20 00:00:00",
            date_to="2026-06-20 23:59:59",
            native_pivot=False,
        )
        workbook = load_workbook(export.output_path)
        self.assertIn("Data", workbook.sheetnames)
        self.assertIn("Summary", workbook.sheetnames)
        summary = workbook["Summary"]
        header = [cell.value for cell in summary[1]]
        self.assertEqual(header, ["txn_result", "count", "percent"])
        total = sum(row[1].value for row in summary.iter_rows(min_row=2))
        self.assertEqual(total, export.row_count)

    def test_single_day_export_includes_champion_sheets(self) -> None:
        from openpyxl import load_workbook

        save_daily_csvs(
            [
                _areq("T1", "2026-06-20 10:00:00.000", acct="4111111111111111"),
                _ares("T1", "2026-06-20 10:00:01.000"),
                _areq("T2", "2026-06-20 10:15:00.000", acct="4111111111111111"),
                _ares("T2", "2026-06-20 10:15:01.000"),
                _areq("T3", "2026-06-20 11:00:00.000", acct="5555555555554444"),
                _ares("T3", "2026-06-20 11:00:01.000"),
            ]
        )
        export = file_report.export_report_xlsx(
            mode="date",
            date_from="2026-06-20 00:00:00",
            date_to="2026-06-20 23:59:59",
            native_pivot=False,
        )
        workbook = load_workbook(export.output_path)
        self.assertIn("Champions", workbook.sheetnames)
        self.assertIn("Champion Attempts", workbook.sheetnames)

        champions = workbook["Champions"]
        champion_header = [cell.value for cell in champions[1]]
        self.assertEqual(champion_header[0], "champion_rank")
        self.assertEqual(champions[2][0].value, 1)
        self.assertEqual(champions[2][2].value, "411111******1111")
        self.assertEqual(champions[2][5].value, 2)
        self.assertEqual(champions[3][0].value, 2)
        self.assertEqual(champions[3][5].value, 1)

        attempts = workbook["Champion Attempts"]
        attempt_header = [cell.value for cell in attempts[1]]
        self.assertEqual(attempt_header[4], "attempt_no")
        self.assertEqual(attempts[2][0].value, 1)
        self.assertEqual(attempts[2][4].value, 1)
        self.assertEqual(attempts[3][0].value, 1)
        self.assertEqual(attempts[3][4].value, 2)
        self.assertEqual(export.row_count, 3)

    def test_multi_day_export_skips_champion_sheets(self) -> None:
        from openpyxl import load_workbook

        self._seed_day("2026-06-20", "T1")
        self._seed_day("2026-06-21", "T2")
        export = file_report.export_report_xlsx(
            mode="date",
            date_from="2026-06-20 00:00:00",
            date_to="2026-06-21 23:59:59",
            native_pivot=False,
        )
        workbook = load_workbook(export.output_path)
        self.assertNotIn("Champions", workbook.sheetnames)
        self.assertNotIn("Champion Attempts", workbook.sheetnames)

    def test_pagination_is_consistent_across_pages(self) -> None:
        save_daily_csvs(
            [
                _areq("T1", "2026-06-20 10:00:00.000"),
                _areq("T2", "2026-06-20 10:01:00.000"),
                _areq("T3", "2026-06-20 10:02:00.000"),
            ]
        )
        kwargs = dict(
            mode="date",
            date_from="2026-06-20 00:00:00",
            date_to="2026-06-20 23:59:59",
        )
        page1 = file_report.run_report_query(limit=2, offset=0, **kwargs)
        page2 = file_report.run_report_query(limit=2, offset=2, **kwargs)
        self.assertEqual(page1.row_count, 3)
        self.assertEqual(page2.row_count, 3)
        self.assertEqual(len(page1.rows), 2)
        self.assertEqual(len(page2.rows), 1)
        seen = {row["threedsservertransid"] for row in page1.rows + page2.rows}
        self.assertEqual(seen, {"T1", "T2", "T3"})

    def test_cache_invalidates_when_data_changes(self) -> None:
        kwargs = dict(
            mode="date",
            date_from="2026-06-20 00:00:00",
            date_to="2026-06-20 23:59:59",
        )
        self._seed_day("2026-06-20", "T1")
        first = file_report.run_report_query(**kwargs)
        self.assertEqual(first.row_count, 1)
        save_daily_csvs(
            [
                _areq("T1", "2026-06-20 10:00:00.000"),
                _areq("T9", "2026-06-20 12:00:00.000"),
            ]
        )
        second = file_report.run_report_query(**kwargs)
        self.assertEqual(second.row_count, 2)

    def test_native_pivot_has_no_row_cap(self) -> None:
        from unittest.mock import patch

        save_daily_csvs(
            [
                _areq("T1", "2026-06-20 10:00:00.000"),
                _areq("T2", "2026-06-20 11:00:00.000"),
            ]
        )
        with patch("app.services.excel_pivot.native_pivot_available", return_value=True), patch(
            "app.services.excel_pivot.add_native_pivot"
        ) as mock_add_native_pivot:
            export = file_report.export_report_xlsx(
                mode="date",
                date_from="2026-06-20 00:00:00",
                date_to="2026-06-20 23:59:59",
                native_pivot=True,
            )
        mock_add_native_pivot.assert_called_once()
        self.assertTrue(export.pivot_added)
        self.assertEqual(export.pivot_error, "")

    def test_native_pivot_reports_when_excel_unavailable(self) -> None:
        from unittest.mock import patch

        save_daily_csvs([_areq("T1", "2026-06-20 10:00:00.000")])
        with patch("app.services.excel_pivot.native_pivot_available", return_value=False):
            export = file_report.export_report_xlsx(
                mode="date",
                date_from="2026-06-20 00:00:00",
                date_to="2026-06-20 23:59:59",
                native_pivot=True,
            )
        self.assertFalse(export.pivot_added)
        self.assertIn("Excel is not available", export.pivot_error)

    def test_native_pivot_warns_when_columns_missing(self) -> None:
        self._seed_day("2026-06-20", "T1")
        export = file_report.export_report_xlsx(
            mode="custom",
            date_from="2026-06-20 00:00:00",
            date_to="2026-06-20 23:59:59",
            sql=(
                "SELECT threedsservertransid, messagedatetime AS areq_messagedatetime "
                "FROM cust_acs_3dsmess WHERE messagetype = 'AReq'"
            ),
            native_pivot=True,
        )
        self.assertFalse(export.pivot_added)
        self.assertIn("missing one or more columns", export.pivot_error)

    def test_resolve_missing_day_lists_all_missing(self) -> None:
        self._seed_day("2026-06-20", "T1")
        with self.assertRaises(ValueError) as ctx:
            resolve_csv_paths_for_dates(["2026-06-20", "2026-06-21", "2026-06-22"])
        message = str(ctx.exception)
        self.assertIn("2026-06-21", message)
        self.assertIn("2026-06-22", message)


if __name__ == "__main__":
    unittest.main()
